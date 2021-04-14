import openpyxl


class Xlsx:
    # Connection
    # readExcelFile = None
    # writeExcelFile = None
    # read_current_row = 2
    # write_current_row = 2
    # read_sheet = None
    # write_sheet = None

    def __init__(self):
        self.readExcelFile = openpyxl.load_workbook('ConvertDBObject.xlsx')
        self.writeExcelFile = openpyxl.load_workbook('fommater_new.xlsx')

        self.read_sheet = self.readExcelFile.active  # 엑셀은 최소한 1개이상의 sheet는 존재. 기본적으로 존재하는 sheet을 가져올때 이렇게.
        self.write_sheet = self.writeExcelFile.active

        self.read_current_row = 2
        self.write_current_row = 2


    def read_xlsx(self):

        while True:
            row_number = self.read_sheet.cell(row=self.read_current_row, column=1).value
            complete_flag = self.read_sheet.cell(row=self.read_current_row, column=4).value

            if row_number is None:
                return {
                    'eof':True
                }
            if row_number is not None and complete_flag is not None:
                self.read_current_row += 1

            else:
                type = self.read_sheet.cell(row=self.read_current_row, column=2).value
                obj_name = self.read_sheet.cell(row=self.read_current_row, column=3).value
                self.read_current_row += 1
                return {
                    'row_number':row_number,
                    'type':type,
                    'obj_name':obj_name,
                    'eof':False
                }

    def save_point(self):
        self.read_sheet.cell(row=(self.read_current_row-1), column=4).value = True

        self.writeExcelFile.save('fommater_new.xlsx')
        self.readExcelFile.save('ConvertDBObject.xlsx')



    def write_xlsx(self, read_param, aa):

        while True:
            row_number = self.write_sheet.cell(row=self.write_current_row, column=1).value
            if row_number is not None:
                self.write_current_row += 1
            else:
                if len(aa) > 0:

                    print(aa)

                    write_sequnce = 1
                    row_number = read_param['row_number']
                    type = read_param['type']
                    obj_name = read_param['obj_name']

                    for a in aa:
                        first_args = a['first_args']
                        second_args = a['second_args']
                        last_args = a['last_args']
                        first_args_result = a['first_args_result']
                        secound_args_result = a['secound_args_result']
                        last_args_result = a['last_args_result']
                        parsing_source = a['source']
                        # first args : CHAR이 포함되었을때 True 반환
                        # True/True 일때 False 일때

                        total_args_result = ((a['first_args_result'] and a['last_args_result']) or (a['first_args_result'] == False) or (a['secound_args_result'] == True))

                        # row_number
                        self.write_sheet.cell(row=self.write_current_row, column=1).value = row_number

                        # WriteNo
                        self.write_sheet.cell(row=self.write_current_row, column=2).value = write_sequnce

                        # Type
                        self.write_sheet.cell(row=self.write_current_row, column=3).value = type
                        # obj_name
                        self.write_sheet.cell(row=self.write_current_row, column=4).value = obj_name

                        # First_args
                        self.write_sheet.cell(row=self.write_current_row, column=5).value = first_args
                        # Second_args
                        self.write_sheet.cell(row=self.write_current_row, column=6).value = second_args
                        # Last_args
                        self.write_sheet.cell(row=self.write_current_row, column=7).value = last_args
                        # First_args_result
                        self.write_sheet.cell(row=self.write_current_row, column=8).value = first_args_result
                        # Last_args_result
                        self.write_sheet.cell(row=self.write_current_row, column=9).value = last_args_result
                        # ID등 포함여부 "ID", "QTY", "QUANTITY", "NO", "SUM"
                        self.write_sheet.cell(row=self.write_current_row, column=10).value = secound_args_result

                        # Total_result
                        self.write_sheet.cell(row=self.write_current_row, column=11).value = total_args_result

                        # 구문
                        self.write_sheet.cell(row=self.write_current_row, column=13).value = parsing_source

                        write_sequnce += 1
                        self.write_current_row += 1

                self.save_point()
                return


